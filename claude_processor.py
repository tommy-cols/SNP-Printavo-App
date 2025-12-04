"""
Claude AI processor for extracting order data from various file formats.
Handles PDFs, images, Excel files, CSV, and text content.
"""

import anthropic
import base64
import os
import csv
import tempfile
import json
from datetime import datetime
from typing import List, Tuple, Dict, Optional

# Try importing optional dependencies
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

# Constants - Use the model that works
CLAUDE_MODEL = "claude-sonnet-4-20250514"  # Stable Sonnet 4
MAX_TOKENS = 16000

# Confidence thresholds
HIGH_CONFIDENCE = 0.90
MEDIUM_CONFIDENCE = 0.70
LOW_CONFIDENCE_THRESHOLD = 0.70


def process_files_with_claude(
        files: List[str],
        text_content: str,
        custom_prompt: str,
        api_key: str,
        status_callback=None
) -> Tuple[bool, Optional[str], Dict]:
    """
    Main function - processes files and text with Claude AI to extract order data.

    Args:
        files: List of file paths to process
        text_content: Direct text/email content to process
        custom_prompt: Custom user prompt (empty string = use default)
        api_key: Claude API key
        status_callback: Optional callback function for status updates

    Returns:
        Tuple of (success: bool, csv_path: str or None, metadata: dict)
    """

    def update_status(message: str):
        """Helper to update status if callback provided"""
        if status_callback:
            status_callback(message)

    try:
        update_status("Building content blocks from files...")

        # Build content blocks for Claude
        content_blocks, processed_files = build_content_blocks(files, text_content)

        if not content_blocks:
            return False, None, {"error": "No valid content to process"}

        update_status(f"Processing {len(processed_files)} files with Claude AI...")

        # Build the prompt
        prompt = build_extraction_prompt(text_content, processed_files, custom_prompt)

        # Insert prompt at the beginning
        content_blocks.insert(0, {"type": "text", "text": prompt})

        update_status("Calling Claude API...")

        # Call Claude with structured output
        result = call_claude_structured_output(content_blocks, api_key)

        if not result or "items" not in result:
            return False, None, {"error": "No items extracted from Claude response"}

        items = result.get("items", [])

        if len(items) == 0:
            return False, None, {"error": "Claude did not extract any order items"}

        update_status(f"Extracted {len(items)} items, saving to CSV...")

        # Check confidence
        overall_conf = result.get("overall_confidence", 0.0)
        low_conf_items = [item for item in items if item.get("confidence", 0.0) < LOW_CONFIDENCE_THRESHOLD]

        if low_conf_items:
            update_status(f"⚠️ Warning: {len(low_conf_items)} items have low confidence")

        # Save to CSV
        csv_path = save_result_to_csv(items)

        metadata = {
            "item_count": len(items),
            "files_processed": len(processed_files),
            "overall_confidence": overall_conf,
            "low_confidence_count": len(low_conf_items),
            "flags": result.get("flags", []),
            "extraction_notes": result.get("extraction_notes", "")
        }

        update_status(f"Success! {len(items)} items extracted ({int(overall_conf * 100)}% confidence).")

        return True, csv_path, metadata

    except anthropic.APIError as e:
        error_msg = f"Claude API error: {str(e)}"
        update_status(error_msg)
        return False, None, {"error": error_msg}
    except Exception as e:
        error_msg = f"Processing error: {str(e)}"
        update_status(error_msg)
        return False, None, {"error": error_msg}


def build_content_blocks(files: List[str], text_content: str) -> Tuple[List[Dict], List[Dict]]:
    """
    Build content array for Claude API from files and text.

    Returns:
        Tuple of (content_blocks, processed_files_info)
    """
    content_blocks = []
    processed_files = []

    # Process each file
    for file_path in files:
        if not os.path.exists(file_path):
            continue

        filename = os.path.basename(file_path)
        _, ext = os.path.splitext(filename.lower())

        try:
            # PDF files
            if ext == '.pdf':
                with open(file_path, 'rb') as f:
                    pdf_bytes = f.read()
                    base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')

                content_blocks.append({
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": base64_pdf
                    }
                })
                processed_files.append({"type": "PDF", "name": filename})

            # Image files
            elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                with open(file_path, 'rb') as f:
                    image_bytes = f.read()
                    base64_image = base64.b64encode(image_bytes).decode('utf-8')

                # Determine media type
                media_type_map = {
                    '.jpg': 'image/jpeg',
                    '.jpeg': 'image/jpeg',
                    '.png': 'image/png',
                    '.gif': 'image/gif',
                    '.webp': 'image/webp'
                }
                media_type = media_type_map.get(ext, 'image/jpeg')

                content_blocks.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": base64_image
                    }
                })
                processed_files.append({"type": "Image", "name": filename})

            # CSV files
            elif ext == '.csv':
                with open(file_path, 'r', encoding='utf-8') as f:
                    csv_text = f.read()

                content_blocks.append({
                    "type": "text",
                    "text": f"\n--- CSV File: {filename} ---\n{csv_text}\n--- End of {filename} ---\n"
                })
                processed_files.append({"type": "CSV", "name": filename})

            # Excel files
            elif ext in ['.xlsx', '.xls']:
                excel_text = extract_excel_as_text(file_path)

                content_blocks.append({
                    "type": "text",
                    "text": f"\n--- Excel File: {filename} ---\n{excel_text}\n--- End of {filename} ---\n"
                })
                processed_files.append({"type": "Excel", "name": filename})

            # Text files
            elif ext in ['.txt', '.eml']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    text = f.read()

                content_blocks.append({
                    "type": "text",
                    "text": f"\n--- Text File: {filename} ---\n{text}\n--- End of {filename} ---\n"
                })
                processed_files.append({"type": "Text", "name": filename})

        except Exception as e:
            # Skip files that fail to process
            print(f"Warning: Could not process {filename}: {e}")
            continue

    # Add pasted text content if provided
    if text_content and text_content.strip():
        content_blocks.append({
            "type": "text",
            "text": f"\n--- Pasted Content ---\n{text_content}\n--- End Pasted Content ---\n"
        })
        processed_files.append({"type": "Text", "name": "Pasted Content"})

    return content_blocks, processed_files


def extract_excel_as_text(file_path: str) -> str:
    """
    Extract Excel file contents as tab-separated text.
    """
    if not HAS_OPENPYXL:
        return f"[Excel file: {os.path.basename(file_path)} - openpyxl not installed]"

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        all_data = ""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            all_data += f"\n━━━━━━ SHEET: {sheet_name} ━━━━━━\n"

            for row in sheet.iter_rows(values_only=True):
                # Convert row to strings, handling None values
                clean_row = [str(cell).strip() if cell is not None else '' for cell in row]
                all_data += '\t'.join(clean_row) + '\n'

            all_data += f"━━━━━━ END SHEET: {sheet_name} ━━━━━━\n"

        return all_data

    except Exception as e:
        return f"[Excel file: {os.path.basename(file_path)} - Extraction failed: {str(e)}]"


def build_extraction_prompt(email_context: str, processed_files: List[Dict], custom_prompt: str) -> str:
    """
    Build the extraction prompt for Claude - MATCHES GOOGLE SCRIPT VERSION.
    """
    # Use custom prompt if provided, otherwise use default
    if custom_prompt and custom_prompt.strip():
        return custom_prompt

    # Build file list
    file_list = "No attachments found" if not processed_files else '\n'.join([f"- {f['type']}: {f['name']}" for f in processed_files])

    # Include email context
    context_section = ""
    if email_context and email_context.strip():
        context_section = f"{email_context}\n"

    prompt = f"""You are an expert order data extraction assistant specialized in garment/apparel orders. Extract ALL line items and assess your confidence in the extraction.

{context_section}
Files attached:
{file_list}

CRITICAL EXTRACTION RULES:

1. **IGNORE MISLEADING COLUMN HEADERS** - Column headers are often wrong or unhelpful. Instead:
   - Look for ACTUAL DATA PATTERNS, not header names
   - Common garment item numbers: alphanumeric codes like LOE722, DT672, PC450, ST350, NKDC2115
   - Item numbers may be embedded in strings like "226-brown" where "226" is the item number
   - Quantities are always numbers (typically 1-999)
   - Sizes: Must be normalized to standard codes (see SIZE NORMALIZATION below)
   - Colors are descriptive text (e.g., "Black", "Navy", "True Royal", "Athletic Gold")
   - Colors may be embedded with item numbers (e.g., "226-brown" means item 226, color brown)

2. **SIZE NORMALIZATION** - Convert ALL sizes to these standard formats: 
    **Valid Adult Sizes:
    ** XS, S, M, L, XL, 2XL, 3XL, 4XL, 5XL, 6XL 
    **Valid Youth Sizes:
    ** YXS, YS, YM, YL, YXL 
    **Valid Baby/Toddler Sizes:
    ** 6M, 12M, 18M, 24M, 2T, 3T, 4T, 5T 
    **Normalization Rules:
    ** - "Small" → "S", "Medium" → "M", "Large" → "L", 
    "Extra Large" → "XL" - "XXL" → "2XL", "XXXL" → "3XL" 
    - "Y-S" or "Youth Small" → "YS" (apply pattern to all youth sizes) 
    - Remove spaces and hyphens: "2 XL" → "2XL", "X-L" → "XL" 
    - **IMPORTANT:** "OSFA", "OS", "One Size", "One Size Fits All" - For one-size items, leave size field EMPTY 
    - Note in confidence: "One-size item"

3. **IGNORE IRRELEVANT COLUMNS** - Skip columns that contain:
   - Session IDs, timestamps, database IDs
   - User names, email addresses
   - Subtotals, tax amounts, monetary values
   - Navigation/UI elements
   - Product descriptions that are just long text (not item numbers)

5. **WHAT TO EXTRACT**:
   - Item Number (style code): Usually 4-10 alphanumeric characters, may be embedded in combined strings
   - Quantity: A number representing how many items
   - Size: Garment size (if present)
   - Color: Color name/description, may need to be extracted from combined strings

5. **MULTIPLE SOURCES**: If data appears in multiple files or formats, extract from ALL sources and COMBINE them into a single list

6. **CONFIDENCE SCORING**:
   For EACH extracted item, provide a confidence score (0.0 to 1.0):
   - 0.90-1.0: Clear, unambiguous data with all fields present
   - 0.70-0.89: Good extraction but some ambiguity (e.g., unclear column, might be missing size)
   - 0.50-0.69: Uncertain extraction (e.g., item number format unusual, color might be wrong)
   - Below 0.50: Very uncertain, likely needs human review

7. **FLAG ISSUES** - Note any problems:
   - "ambiguous_columns" - Column headers don't match content
   - "missing_fields" - Some required data not found
   - "format_unusual" - Data in unexpected format
   - "multiple_interpretations" - Could be interpreted different ways
   - "partial_data" - Only some fields extracted with confidence
   - "combined_fields" - Had to split combined data (e.g., "226-brown")

EXAMPLE OF MISLEADING CSV:
If you see headers like: SessionID, Name, SubTotal, TaxTotal, Total, Qty Value, Value24, ItemNum, ProductName
Ignore: SessionID, Name, SubTotal, TaxTotal, Total, Value24
Focus on: Qty Value (actual quantity), ItemNum (item number), ProductName (might contain color/size)

RESPONSE FORMAT:
You must return valid JSON with this structure:
{{
  "items": [
    {{
      "qty": 1,
      "size": "M",
      "item_num": "LOE722",
      "color": "Propel Navy",
      "confidence": 0.95,
      "notes": "Clear extraction from well-structured table"
    }}
  ],
  "overall_confidence": 0.88,
  "flags": ["ambiguous_columns"],
  "extraction_notes": "Column headers were misleading but found clear item data in rows. Combined data from 2 sources."
}}

Extract ALL items from ALL sources and provide confidence scoring:"""

    return prompt


def call_claude_structured_output(content_blocks: List[Dict], api_key: str) -> Optional[Dict]:
    """
    Call Claude API and parse JSON response.
    Note: Structured outputs not available on this model, relying on prompt quality.
    """
    # Initialize Claude client
    client = anthropic.Anthropic(api_key=api_key)

    # Make standard API call - no structured outputs since model doesn't support it
    response = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=MAX_TOKENS,
        temperature=0.2,  # Lower temperature for more consistent extraction
        messages=[{
            "role": "user",
            "content": content_blocks
        }]
    )

    # Debug: Print full response structure
    print(f"\n=== DEBUG: Full Response ===")
    print(f"Response object type: {type(response)}")
    print(f"Response content length: {len(response.content)}")

    # Extract text from response
    response_text = ""
    for idx, content_block in enumerate(response.content):
        print(f"Content block {idx}: type={content_block.type}")
        if content_block.type == "text":
            response_text += content_block.text

    print(f"\n=== Raw Response Text (first 1000 chars) ===")
    print(response_text[:1000])
    print(f"=== End Raw Response ===\n")

    # Check if response is empty
    if not response_text or not response_text.strip():
        raise ValueError("Claude returned an empty response. Check your API key and model access.")

    # Clean up any markdown formatting and preamble text
    response_text = response_text.strip()

    # Find JSON block - it might have text before it
    json_start = response_text.find('{')
    json_end = response_text.rfind('}')

    if json_start == -1 or json_end == -1:
        # Try to find it within code blocks
        if '```json' in response_text:
            json_start = response_text.find('```json') + 7
            json_end = response_text.find('```', json_start)
            response_text = response_text[json_start:json_end].strip()
        elif '```' in response_text:
            json_start = response_text.find('```') + 3
            json_end = response_text.find('```', json_start)
            response_text = response_text[json_start:json_end].strip()
        else:
            raise ValueError("Could not find JSON object in response")
    else:
        # Extract just the JSON object
        response_text = response_text[json_start:json_end + 1]

    response_text = response_text.strip()

    try:
        result = json.loads(response_text)

        # Validate that we got the expected structure
        if "items" not in result:
            raise ValueError("Response missing 'items' field")

        # Add defaults for optional fields if missing
        if "overall_confidence" not in result:
            result["overall_confidence"] = 0.8
        if "flags" not in result:
            result["flags"] = []
        if "extraction_notes" not in result:
            result["extraction_notes"] = "Extraction completed"

        # Ensure each item has confidence
        for item in result["items"]:
            if "confidence" not in item:
                item["confidence"] = 0.8
            if "notes" not in item:
                item["notes"] = ""

        print(f"✓ Successfully parsed JSON with {len(result['items'])} items")
        return result

    except json.JSONDecodeError as e:
        print(f"\n=== JSON PARSE ERROR ===")
        print(f"Error: {e}")
        print(f"Response text length: {len(response_text)}")
        print(f"First 1000 chars of cleaned response:")
        print(response_text[:1000])
        print(f"Last 200 chars of cleaned response:")
        print(response_text[-200:] if len(response_text) > 200 else response_text)
        raise ValueError(f"Claude did not return valid JSON: {str(e)}")


def save_result_to_csv(items: List[Dict]) -> str:
    """
    Save extraction results to CSV file in temp directory.
    Include confidence scores like Google Script version.

    Returns:
        Path to saved CSV file
    """
    # Create temp directory for app
    app_temp_dir = os.path.join(tempfile.gettempdir(), 'printavo_quote_creator')
    os.makedirs(app_temp_dir, exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(app_temp_dir, f'claude_output_{timestamp}.csv')

    # Write CSV with confidence column
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        # Include confidence as last column (like Google Script)
        fieldnames = ['qty', 'size', 'item num', 'color', 'confidence']
        writer = csv.DictWriter(f, fieldnames=fieldnames)

        writer.writeheader()

        for item in items:
            confidence_pct = f"{int(item.get('confidence', 0) * 100)}%"
            writer.writerow({
                'qty': item.get('qty', ''),
                'size': item.get('size', ''),
                'item num': item.get('item_num', ''),
                'color': item.get('color', ''),
                'confidence': confidence_pct
            })

    return csv_path